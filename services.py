"""
Сервисные функции для бота управления парковкой
"""

from datetime import datetime
from typing import Optional, List, Tuple
from io import BytesIO

from sqlalchemy.orm import Session
from aiogram.types import Message
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from models import (
    User, Role as RoleModel, Parking, Task, ParkingQueue, Break
)
from vehicle_validator import VehicleNumberValidator
from utils import (
    ensure_timezone_aware, get_timezone_aware_now, format_duration, get_current_shift_period
)
from config import config


async def get_user(db: Session, telegram_id: int) -> Optional[User]:
    """
    Получение пользователя из базы данных по Telegram ID

    Args:
        db: Сессия базы данных
        telegram_id: Telegram ID пользователя

    Returns:
        Объект User или None
    """
    return db.query(User).filter(User.telegram_id == telegram_id).first()


async def get_or_create_user(db: Session, message: Message) -> User:
    """
    Получение существующего или создание нового пользователя

    Args:
        db: Сессия базы данных
        message: Сообщение от пользователя

    Returns:
        Объект User
    """
    user = await get_user(db, message.from_user.id)

    if not user:
        # Создание нового пользователя с базовой ролью DRIVER
        user = User(
            telegram_id=message.from_user.id,
            username=message.from_user.username,
            first_name=message.from_user.first_name or "",
            last_name=message.from_user.last_name or "",
            current_role="DRIVER",
            created_at=get_timezone_aware_now(),
            is_on_shift=False,
            is_on_break=False,
            total_break_time=0
        )
        db.add(user)
        db.flush()

        # Добавление роли DRIVER
        driver_role = db.query(RoleModel).filter(RoleModel.name == "DRIVER").first()
        if driver_role:
            user.roles.append(driver_role)

        db.commit()
        db.refresh(user)

    return user


def get_user_roles(user: User) -> List[str]:
    """Получение списка названий ролей пользователя"""
    return [role.name for role in user.roles]


def get_user_main_role(user: User) -> str:
    """
    Определение основной роли пользователя для отображения меню

    Приоритет ролей:
    1. Текущая выбранная роль (если есть и существует)
    2. ADMIN
    3. OPERATOR
    4. DRIVER_TRANSFER
    5. DEB_EMPLOYEE
    6. DRIVER (по умолчанию)

    Returns:
        Название роли
    """
    user_roles = get_user_roles(user)

    # Если нет ролей - даем DRIVER
    if not user_roles:
        return "DRIVER"

    # Приоритет текущей роли
    if user.current_role and user.current_role in user_roles:
        return user.current_role

    # Приоритет ролей
    role_priority = ["ADMIN", "OPERATOR", "DRIVER_TRANSFER", "DEB_EMPLOYEE", "DRIVER"]
    for role in role_priority:
        if role in user_roles:
            return role

    return "DRIVER"


async def add_to_parking_queue(db: Session, user_id: int, vehicle_number: str, is_hitch: bool) -> ParkingQueue:
    """
    Добавление ТС в очередь на парковку
    """
    vehicle_type = "HITCH" if is_hitch else "NON_HITCH"

    queue_item = ParkingQueue(
        user_id=user_id,
        vehicle_number=vehicle_number,
        is_hitch=is_hitch,
        vehicle_type=vehicle_type,
        created_at=get_timezone_aware_now(),
        status="waiting"
    )
    db.add(queue_item)
    db.commit()
    db.refresh(queue_item)
    return queue_item


async def get_queue_position(db: Session, user_id: int) -> int:
    """
    Получение позиции в очереди
    """
    user_item = db.query(ParkingQueue).filter(
        ParkingQueue.user_id == user_id,
        ParkingQueue.status.in_(["waiting", "notified"])
    ).first()

    if not user_item:
        return 0

    position = db.query(ParkingQueue).filter(
        ParkingQueue.status.in_(["waiting", "notified"]),
        ParkingQueue.created_at < user_item.created_at
    ).count() + 1

    return position


async def get_queue_stats(db: Session) -> dict:
    """
    Получение статистики очереди
    """
    waiting = db.query(ParkingQueue).filter(
        ParkingQueue.status == "waiting"
    ).count()

    notified = db.query(ParkingQueue).filter(
        ParkingQueue.status == "notified"
    ).count()

    total = waiting + notified

    hitch_count = db.query(ParkingQueue).filter(
        ParkingQueue.status.in_(["waiting", "notified"]),
        ParkingQueue.is_hitch == True
    ).count()

    non_hitch_count = total - hitch_count

    return {
        "total": total,
        "waiting": waiting,
        "notified": notified,
        "hitch": hitch_count,
        "non_hitch": non_hitch_count
    }


async def process_parking_departure(db: Session, spot_number: int, bot):
    """
    Обработка освобождения парковочного места
    - Отправка уведомления следующему в очереди
    """
    # Находим следующего в очереди
    next_in_queue = db.query(ParkingQueue).filter(
        ParkingQueue.status == "waiting"
    ).order_by(ParkingQueue.created_at.asc()).first()

    if next_in_queue:
        next_in_queue.status = "notified"
        next_in_queue.notified_at = get_timezone_aware_now()
        next_in_queue.spot_number = spot_number
        db.commit()

        # Отправляем уведомление
        user = await get_user(db, next_in_queue.user_id)
        if user:
            try:
                from utils import Emoji
                await bot.send_message(
                    user.telegram_id,
                    f"{Emoji.NOTIFIED} ОСВОБОДИЛОСЬ ПАРКОВОЧНОЕ МЕСТО!\n\n"
                    f"📍 Место #{spot_number} свободно\n"
                    f"🚗 Ваше ТС: {next_in_queue.vehicle_number}\n"
                    f"📝 Тип: {'Перецепной' if next_in_queue.is_hitch else 'Не перецепной'}\n"
                    f"⏰ Время ожидания: {format_duration(int((get_timezone_aware_now() - ensure_timezone_aware(next_in_queue.created_at)).total_seconds()))}\n"
                    f"📊 Позиция в очереди: 1\n\n"
                    f"{Emoji.DRIVE} Проследуйте на парковочное место #{spot_number}\n"
                    f"{Emoji.ARRIVED} После парковки нажмите кнопку 'Прибытие'"
                )
            except Exception as e:
                print(f"Ошибка отправки уведомления: {e}")

        return next_in_queue
    return None


async def get_free_parking_spot(db: Session) -> Optional[int]:
    """
    Поиск свободного парковочного места

    Args:
        db: Сессия базы данных

    Returns:
        Номер свободного места или None
    """
    total_spots = config.PARKING_SPOTS
    occupied_spots = db.query(Parking).filter(
        Parking.departure_time == None
    ).all()

    occupied_numbers = {spot.spot_number for spot in occupied_spots}

    if len(occupied_numbers) >= total_spots:
        return None

    for spot_num in range(1, total_spots + 1):
        if spot_num not in occupied_numbers:
            return spot_num
    return None


async def validate_vehicle_number(vehicle_number: str) -> bool:
    """Валидация номера ТС"""
    is_valid, _ = VehicleNumberValidator.validate(vehicle_number)
    return is_valid


async def validate_vehicle_number_with_explanation(vehicle_number: str) -> Tuple[bool, str]:
    """Валидация номера ТС с пояснением ошибки"""
    return VehicleNumberValidator.validate(vehicle_number)


async def normalize_vehicle_number(vehicle_number: str) -> str:
    """Нормализация номера ТС (латиница -> кириллица)"""
    return VehicleNumberValidator.normalize(vehicle_number)


async def get_active_transfer_drivers(db: Session) -> List[User]:
    """
    Получение всех активных водителей перегона на смене

    Args:
        db: Сессия базы данных

    Returns:
        Список водителей перегона на смене
    """
    return db.query(User).filter(
        User.is_on_shift == True,
        User.is_on_break == False,
        User.roles.any(RoleModel.name == "DRIVER_TRANSFER")
    ).all()


async def get_task_from_pool(db: Session) -> Optional[Task]:
    """
    Получение задачи с максимальным приоритетом из пула
    - Проверяет, что ТС еще на парковке
    - Проверяет, что задача в статусе PENDING
    - Проверяет, что задача в пуле
    - Сортирует по приоритету и времени создания
    """
    return db.query(Task).join(Parking).filter(
        Task.status == "PENDING",
        Task.is_in_pool == True,
        Parking.is_hitch == True,
        Parking.departure_time == None,  # ТС еще на парковке
        Task.driver_id == None  # Никто не взял
    ).order_by(
        Task.priority.desc(),
        Task.created_at.asc()
    ).first()


async def generate_excel_report(parkings: List[Parking], period: str = None) -> BytesIO:
    """
    Генерация Excel отчета по парковке

    Args:
        parkings: Список записей парковки
        period: Название периода

    Returns:
        BytesIO объект с Excel файлом
    """
    wb = Workbook()
    ws = wb.active
    ws.title = f"Отчет {period}" if period else "Отчет парковки"

    # Заголовки
    headers = [
        "№", "Тип ТС", "Номер ТС", "Номер парковочного места",
        "Дата и время прибытия", "Дата и время убытия",
        "Время на парковке", "Время убытия из отделения", "Время в отделении"
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Данные
    for idx, parking in enumerate(parkings, 2):
        ws.cell(row=idx, column=1, value=idx-1)
        ws.cell(row=idx, column=2, value="Перецепной" if parking.is_hitch else "Не перецепной")
        ws.cell(row=idx, column=3, value=parking.vehicle_number)
        ws.cell(row=idx, column=4, value=parking.spot_number)

        if parking.arrival_time:
            ws.cell(row=idx, column=5, value=parking.arrival_time.strftime("%d.%m.%Y %H:%M:%S"))
        if parking.departure_time:
            ws.cell(row=idx, column=6, value=parking.departure_time.strftime("%d.%m.%Y %H:%M:%S"))

    # Ширина колонок
    column_widths = [5, 15, 15, 20, 20, 20, 15, 25, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    return excel_file

# Добавьте в конец файла services.py

def get_tasks_for_current_shift(db: Session, include_completed: bool = False) -> List[Task]:
    """
    Получение задач за текущую смену
    """
    start_time, end_time, _ = get_current_shift_period()

    query = db.query(Task).filter(
        Task.created_at >= start_time,
        Task.created_at <= end_time
    )

    if not include_completed:
        # Исключаем выполненные задачи, если не запрошены специально
        query = query.filter(Task.status != "COMPLETED")

    return query.order_by(Task.created_at.desc()).all()


def get_parking_for_current_shift(db: Session) -> List[Parking]:
    """
    Получение записей о парковке за текущую смену
    Включает только активные (без departure_time) и те, что были созданы в эту смену
    """
    start_time, end_time, _ = get_current_shift_period()

    # Активные парковки (ТС еще на месте)
    active_parkings = db.query(Parking).filter(
        Parking.departure_time == None
    ).all()

    # Парковки, созданные в эту смену (даже если уже убыли)
    shift_parkings = db.query(Parking).filter(
        Parking.arrival_time >= start_time,
        Parking.arrival_time <= end_time
    ).all()

    # Объединяем без дубликатов
    parking_dict = {p.id: p for p in active_parkings}
    for p in shift_parkings:
        parking_dict[p.id] = p

    return list(parking_dict.values())


def get_queue_for_current_shift(db: Session) -> List[ParkingQueue]:
    """
    Получение записей очереди за текущую смену
    """
    start_time, end_time, _ = get_current_shift_period()

    # Активные в очереди
    active_queue = db.query(ParkingQueue).filter(
        ParkingQueue.status.in_(["waiting", "notified"])
    ).all()

    # Записи, созданные в эту смену
    shift_queue = db.query(ParkingQueue).filter(
        ParkingQueue.created_at >= start_time,
        ParkingQueue.created_at <= end_time
    ).all()

    # Объединяем без дубликатов
    queue_dict = {q.id: q for q in active_queue}
    for q in shift_queue:
        queue_dict[q.id] = q

    return list(queue_dict.values())
